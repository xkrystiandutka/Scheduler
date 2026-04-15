import random, calendar
from datetime import datetime, timedelta, date
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from .utils import month_days, polish_holidays
import os

class Scheduler:
    def __init__(self, seed=None):
        if seed is not None:
            random.seed(seed)

        # 1. Definiujemy grupy
        self.special_rotation_2 = ["T Marek", "N Wojciech", "K Hubert"]
        self.special_rotation = ["P Barbara", "F Beata", "P Jacek", "D Krystian"]
        
        # 2. Definiujemy "zwykłych" pracowników
        normal_staff = [
            "F Tomasz", "F Krzysztof", "S Sławomir", "G Artur",
            "M Zbigniew", "W Magdalena", "D Janusz"
        ]

        # 3. Łączymy w poprawnej kolejności: Zwykli -> Grupa 2 -> Grupa 1
        # Dzięki temu w Excelu Marek będzie pod Januszem, ale przed Barbarą.
        self.employees = normal_staff + self.special_rotation_2 + self.special_rotation

        # ... reszta słowników bez zmian ...
        self.SHIFTS = {
            "07.00-15.00": (7, 15, 8), "14.00-22.00": (14, 22, 8), "08.00-17.00": (8, 17, 9),
            "07.00-14.00": (7, 14, 7), "07.00-13.00": (7, 13, 6), "07.00-12.00": (7, 12, 5), 
            "13.00-21.00": (13, 21, 8), "08.00-16.00": (8, 16, 8),
            "14.00-21.00": (14, 21, 7), "14.00-20.00": (14, 20, 6), "14.00-19.00": (14, 19, 5), 
            "12.00-20.00": (12, 20, 8),
            "OFF": (None, None, 0), "WN": (None, None, 0), "WS": (None, None, 0),
            "WP": (None, None, 0), "WW": (None, None, 0), "WH": (None, None, 0),
        }
        self.COLORS = {
            "saturday": "FF892E", "sunday": "FF892E", "holiday": "CD3C32", "header": "CD3C32", "odbior": "92D050"
        }

    def shift_times(self, code, dt):
        if code in ("OFF", "WN", "WS", "WP", "WH", "WW") or code not in self.SHIFTS:
            return None
        sh, eh, hrs = self.SHIFTS[code]
        start = datetime(dt.year, dt.month, dt.day, sh)
        end = datetime(dt.year, dt.month, dt.day, eh)
        if eh <= sh: end += timedelta(days=1)
        return start, end

    def rest_ok(self, prev_code, prev_date, next_code, next_date):
        if prev_code in (None, "OFF", "WN", "WS", "WP", "WH") or next_code in (None, "OFF", "WN", "WS", "WP", "WH"):
            return True
        p_t = self.shift_times(prev_code, prev_date)
        n_t = self.shift_times(next_code, next_date)
        if not p_t or not n_t: return True
        return (n_t[0] - p_t[1]).total_seconds() / 3600.0 >= 16

    def week_index(self, d):
        first = d.replace(day=1)
        first_monday = first - timedelta(days=first.weekday())
        return (d - first_monday).days // 7

    def _make_weekly_pref(self, weeks, employees):
        weekly_pref = {}
        
        rotation_1 = self.special_rotation 
        rotation_2 = self.special_rotation_2 
        
        all_special = set(rotation_1) | set(rotation_2)
        normal_candidates = [e for e in employees if e not in all_special]
        random.shuffle(normal_candidates)

        def get_next_afternoon_worker():
            idx = 0
            while True:
                yield normal_candidates[idx % len(normal_candidates)]
                idx += 1

        afternoon_gen = get_next_afternoon_worker()
        sorted_weeks = sorted(weeks)

        # Sprawdzamy długość pierwszego tygodnia
        first_week_days = [d for d in self.days if self.week_index(d) == sorted_weeks[0]]
        work_days_in_first_week = len([d for d in first_week_days if d.weekday() < 5])

        for i, w in enumerate(sorted_weeks):
            weekly_pref[w] = {}
            
            # Logika "krótkiego tygodnia"
            if work_days_in_first_week < 3:
                rot_idx = (i - 1) if i > 0 else 0
            else:
                rot_idx = i

            # --- 1. PRZYPISANIE GRUPY 2 (MAREK I INNI) - PIERWSZA KOLEJNOŚĆ ---
            special_2_person = rotation_2[rot_idx % len(rotation_2)]
            for e in rotation_2:
                if e in employees:
                    # 13-21 dla wybranego, reszta 08-16
                    weekly_pref[w][e] = "13.00-21.00" if e == special_2_person else "08.00-16.00"

            # --- 2. PRZYPISANIE GRUPY 1 (BARBARA I INNI) ---
            special_1_person = rotation_1[rot_idx % len(rotation_1)]
            for e in rotation_1:
                if e in employees:
                    weekly_pref[w][e] = "14.00-22.00" if e == special_1_person else "07.00-15.00"

            # --- 3. PRZYPISANIE GRUPY NORMALNEJ (TOMASZ I INNI) ---
            worker_12_20 = next(afternoon_gen)
            for e in normal_candidates:
                if e in employees:
                    weekly_pref[w][e] = "12.00-20.00" if e == worker_12_20 else "07.00-15.00"
                
        return weekly_pref

    def _assign_weekend_day(self, d, weekly_pref, week_of, schedule, hours, stats, 
                            last_sun_day, last_hol_day, last_sat_day, assigned_today):
        # 1. Definiujemy osoby, które MAJĄ ZAKAZ pracy w weekendy i święta
        forbidden_employees = set(self.special_rotation_2) | set(self.special_rotation)
        
        weekday = d.weekday()
        is_hol = d in polish_holidays(d.year)
        
        # Ustalamy numer tygodnia w miesiącu, aby wiedzieć czy dyżur jest 1- czy 2-osobowy
        sorted_weeks = sorted(set(week_of.values()))
        nth_week = sorted_weeks.index(week_of[d]) + 1
        
        # Logika obsady: Święta = 2 osoby, weekendy co drugi tydzień 2 osoby, inaczej 1 osoba
        if is_hol:
            num_workers = 2
            forced_shift = None
        else:
            num_workers = 1 if nth_week % 2 != 0 else 2
            forced_shift = "08.00-17.00" if (nth_week % 2 != 0) else None

        scored_candidates = []
        for e in schedule.keys():
            # BLOKADA: Jeśli pracownik jest w grupie specjalnej 1 lub 2, pomiń go
            if e in forbidden_employees:
                continue
            
            # Jeśli pracownik ma już przypisane coś na dziś (np. urlop WW)
            if schedule[e].get(d) == "WW" or e in assigned_today:
                continue
            
            # Blokada: Nie pracujemy w niedzielę, jeśli była pracująca sobota (zasada odpoczynku)
            if weekday == 6:
                yesterday = d - timedelta(days=1)
                if schedule[e].get(yesterday) not in ("OFF", "WN", "WS", "WP", "WH"):
                    continue

            # Wybór odpowiedniej pamięci i limitów dla punktacji
            if is_hol:
                last = last_hol_day.get(e)
                limit = 10 
                work_count = stats[e]["holidays"]
            elif weekday == 6:
                last = last_sun_day.get(e)
                limit = 21 # Minimum 3 tygodnie odstępu dla niedziel
                work_count = stats[e]["sundays"]
            else: # Sobota
                last = last_sat_day.get(e)
                limit = 12
                work_count = stats[e]["saturdays"]
                
            days_since = (d - last).days if last else 999
            
            # Punktacja (score): 
            # (0, ...) - osoby, które odpoczywały powyżej limitu (priorytet)
            # (1, ...) - osoby, które muszą wejść w kolejkę ratunkową
            if days_since < limit:
                score = (1, -days_since, work_count)
            else:
                score = (0, work_count, -days_since)

            scored_candidates.append({"name": e, "score": score})

        # Sortowanie kandydatów według punktacji i wybór najlepszych
        scored_candidates.sort(key=lambda x: x["score"])
        picked = [c["name"] for c in scored_candidates[:num_workers]]
        
        # Przypisywanie zmian wybranym osobom
        for i, e in enumerate(picked):
            # Ustalanie kodu zmiany (08-17 lub rano/popołudnie)
            shift = forced_shift if forced_shift else ("07.00-15.00" if i == 0 else "14.00-22.00")
            
            # Korekta dla świąt (nie używamy 08-17 w święta)
            if is_hol and shift == "08.00-17.00": 
                shift = "07.00-15.00"

            schedule[e][d] = shift
            hours[e] += self.SHIFTS[shift][2]
            
            # Aktualizacja statystyk i dat ostatniej pracy
            if is_hol:
                stats[e]["holidays"] += 1
                last_hol_day[e] = d
            elif weekday == 6:
                stats[e]["sundays"] += 1
                last_sun_day[e] = d
            elif weekday == 5:
                stats[e]["saturdays"] += 1
                last_sat_day[e] = d
                
            assigned_today.add(e)

    def _assign_weekday(self, d, weekly_pref, week_of, schedule, hours, stats):
        w = week_of[d]
        prev = d - timedelta(days=1)
        
        # 1. Najpierw definiujemy listę pracowników na popołudnie
        pm_workers = [e for e in schedule.keys() if weekly_pref[w][e] in ("14.00-22.00", "12.00-20.00", "13.00-21.00")]        
        
        # 2. Teraz sprawdzamy urlopy i przypisujemy zmiany
        for e in pm_workers:
            if schedule[e].get(d) == "WW": 
                continue # Pomiń jeśli ma urlop
                
            pref = weekly_pref[w][e]
            if self.rest_ok(schedule[e].get(prev, "OFF"), prev, pref, d):
                schedule[e][d] = pref
                hours[e] += 8

        # 2. Reszta (w tym ci po niedzieli)
        for e in schedule.keys():
            if schedule[e][d] != "OFF": continue
            
            target = weekly_pref[w][e]
            prev_shift = schedule[e].get(prev, "OFF")

            if self.rest_ok(prev_shift, prev, target, d):
                schedule[e][d] = target
                hours[e] += 8
            else:
                # Jeśli po weekendzie nie może przyjść rano, wymuszamy popołudnie.
                # Zamiast tracić dzień (WN), wstawiamy go na 14.00-22.00 lub 12.00-20.00.
                if self.rest_ok(prev_shift, prev, "14.00-22.00", d):
                    schedule[e][d] = "14.00-22.00"
                    hours[e] += 8
                else:
                    schedule[e][d] = "12.00-20.00"
                    hours[e] += 8

    def _assign_compensatory(self, employees, days, schedule, hours, week_of, last_sunhol_day):
        """Poprawione odbiory: nie zabierają dni roboczych, jeśli ktoś ma mało godzin."""
        holidays = set(polish_holidays(days[0].year))
        workdays = [d for d in days if d.weekday() < 5 and d not in holidays]
        day_load = Counter()

        # Sortujemy pracowników tak, by ci z największą liczbą godzin pierwsi dostawali odbiory
        sorted_emp = sorted(employees, key=lambda x: hours[x], reverse=True)

        for e in sorted_emp:
            for d in days:
                # Jeśli w dany dzień pracownik ma WW, to nie wypracował w nim odbioru!
                if schedule[e].get(d) == "WW": continue

                shift = schedule[e].get(d)
                if shift not in ("07.00-15.00", "14.00-22.00", "08.00-17.00"): continue
                
                comp = "WN" if d.weekday() == 6 else ("WS" if d in holidays else ("WP" if d.weekday() == 5 else None))
                if not comp: continue

                # Szukamy dnia do odbioru (musi mieć wpisaną zmianę roboczą 07.00-15.00 lub 14.00-22.00)
                possible = [wd for wd in workdays if abs((wd - d).days) <= 7 
                            and schedule[e][wd] in ("07.00-15.00", "14.00-22.00")
                            and schedule[e][wd] != "WW"] # Nie zabieraj dnia, który już jest urlopem!
                
                if possible:
                    # Wybieramy dzień tak, by nie było za dużo odbiorów naraz w biurze
                    possible.sort(key=lambda wd: (day_load[wd], random.random()))
                    cd = possible[0]
                    # Zamieniamy pracę na odbiór (godziny spadają)
                    old_shift_hrs = self.SHIFTS[schedule[e][cd]][2]
                    schedule[e][cd] = comp
                    hours[e] -= old_shift_hrs
                    day_load[cd] += 1

    def _adjust_last_day_hours(self, days, schedule, hours, target_hours=160):
        for e in schedule:
            diff = target_hours - hours[e]
            if diff == 0: continue
            
            # Szukamy ostatniego dnia roboczego (gdzie jest zmiana z "-" np. 07.00-15.00)
            for d in reversed(days):
                current = schedule[e][d]
                # Sprawdzamy czy to dzień roboczy i czy ma w nazwie kreskę (kod zmiany)
                if d.weekday() < 5 and d not in polish_holidays(d.year) and "-" in current:
                    try:
                        # split('-')[0] daje nam "14.00"
                        # split('.')[0] wyciąga z tego samo "14"
                        sh_str = current.split('-')[0].split('.')[0]
                        sh = int(sh_str)
                        
                        current_hrs = self.SHIFTS[current][2]
                        new_hrs = current_hrs + diff
                        
                        if 0 < new_hrs <= 8:
                            new_eh = sh + new_hrs
                            # Tworzymy nowy kod w Twoim formacie: "14.00-20.00"
                            new_code = f"{sh:02d}.00-{int(new_eh):02d}.00"
                            
                            if new_code in self.SHIFTS:
                                schedule[e][d] = new_code
                                hours[e] += diff
                                break
                    except (ValueError, IndexError):
                        continue # Jeśli coś pójdzie nie tak z formatem, szukaj innego dnia

    def generate(self, year, month, employees=None, initial_stats=None, last_weekend_workers=None, leaves=None):
        self.days = month_days(year, month)
        if employees is None:
                    employees = self.employees 
        holidays = set(polish_holidays(year))
        schedule = {e: {d: "OFF" for d in self.days} for e in employees}

        # --- NOWA LOGIKA: WPISYWANIE URLOPÓW NA START ---
        if leaves:
            for emp_name, days_off in leaves.items():
                if emp_name in schedule:
                    for d_num in days_off:
                        # Znajdujemy konkretną datę w self.days
                        target_date = date(year, month, d_num)
                        if target_date in schedule[emp_name]:
                            schedule[emp_name][target_date] = "WW"

        hours = {e: 0 for e in employees}
        week_of = {d: self.week_index(d) for d in self.days}
        
        stats = {e: (initial_stats[e].copy() if initial_stats and e in initial_stats else {"saturdays":0, "sundays":0, "holidays":0}) for e in employees}

        # --- 1. INICJALIZACJA TRZECH OSOBNYCH KOLEJEK ---
        # Data "daleka" (40 dni wstecz), żeby system nie blokował nikogo na starcie bez powodu
        far_past = self.days[0] - timedelta(days=40)
        # Data "niedzielna" (1 dzień wstecz), żeby zablokować konkretne osoby
        last_sunday_of_april = self.days[0] - timedelta(days=1) 

        last_sun_day = {}
        for e in employees:
            if last_weekend_workers and e in last_weekend_workers:
                # Jeśli był w ostatni weekend, wpisujemy mu ostatnią niedzielę kwietnia
                # To go zablokuje w niedzielach na 21 dni od tej daty
                last_sun_day[e] = last_sunday_of_april
            else:
                last_sun_day[e] = far_past

        # Święta i soboty startują z "czystym kontem" dla wszystkich
        last_hol_day = {e: far_past for e in employees}
        last_sat_day = {e: far_past for e in employees}

        weekly_pref = self._make_weekly_pref(set(week_of.values()), employees)

        # --- 2. GENEROWANIE GRAFIKU (WEEKENDY I ŚWIĘTA) ---
        assigned_today = defaultdict(set)
        for d in self.days:
            if d.weekday() in (5, 6) or d in holidays:
                # Przekazujemy wszystkie 3 słowniki do funkcji przypisującej
                self._assign_weekend_day(
                    d, weekly_pref, week_of, schedule, hours, stats, 
                    last_sun_day, last_hol_day, last_sat_day, assigned_today[d]
                )
        
        # --- 3. DNI ROBOCZE I ODBIORY ---
        for d in self.days:
            if d.weekday() < 5 and d not in holidays: 
                self._assign_weekday(d, weekly_pref, week_of, schedule, hours, stats)
        
        # Odbiorami zajmujemy się na końcu (używamy last_sun_day jako bazy)
        self._assign_compensatory(employees, self.days, schedule, hours, week_of, last_sun_day)
        self._adjust_last_day_hours(self.days, schedule, hours)

        # --- 4. PODSUMOWANIE ---
        summary = []
        for e in employees:
            act_h = sum(self.SHIFTS.get(schedule[e][d], (0,0,0))[2] for d in self.days if schedule[e][d] not in ("WN","WS","WP","WH","OFF","WW"))
            summary.append({
                "employee": e, 
                "hours": act_h, 
                "saturdays": stats[e]["saturdays"], 
                "sundays": stats[e]["sundays"], 
                "holidays": stats[e]["holidays"]
            })
            
        return schedule, summary, holidays

    def save_xlsx(self, schedule, summary, holidays, year, month, filename):
        wb = Workbook(); ws = wb.active; ws.title = f"{calendar.month_name[month]}_{year}"
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center"); bold = Font(bold=True)
        sat_f = PatternFill("solid", fgColor=self.COLORS["saturday"]); sun_f = PatternFill("solid", fgColor=self.COLORS["sunday"])
        hol_f = PatternFill("solid", fgColor=self.COLORS["holiday"]); odb_f = PatternFill("solid", fgColor=self.COLORS["odbior"])

        ws.cell(1, 1, "Pracownik").font = bold; ws.cell(1, 2, "Typ danych").font = bold
        for i, d in enumerate(self.days):
            c = 3 + i; ws.cell(1, c, d.day).font = bold
            ws.cell(2, c, calendar.day_name[d.weekday()][:2]).font = bold
            if d in holidays: ws.cell(1, c).fill = hol_f; ws.cell(2, c).fill = hol_f
            elif d.weekday() == 5: ws.cell(1, c).fill = sat_f; ws.cell(2, c).fill = sat_f
            elif d.weekday() == 6: ws.cell(1, c).fill = sun_f; ws.cell(2, c).fill = sun_f

        row = 3
        for e in schedule:
            ws.cell(row, 1, e).font = bold; ws.cell(row, 2, "Godziny"); ws.cell(row+1, 2, "Liczba h")
            for i, d in enumerate(self.days):
                c = 3 + i; val = schedule[e][d]
                cell_code = ws.cell(row, c, val); cell_code.alignment = center; cell_code.border = border
                if val in ("WN", "WS", "WP", "WH", "WW"): cell_code.fill = odb_f
                elif d in holidays: cell_code.fill = hol_f
                elif d.weekday() == 5: cell_code.fill = sat_f
                elif d.weekday() == 6: cell_code.fill = sun_f
                
                h_val = self.SHIFTS.get(val, (0,0,0))[2] if val not in ("WN","WS","WP","WH","OFF") else 0
                cell_h = ws.cell(row+1, c, h_val); cell_h.alignment = center; cell_h.border = border
            row += 2

        row += 1; ws.cell(row, 1, "PODSUMOWANIE").font = bold; row += 1
        ws.cell(row, 1, "Pracownik").font = bold; ws.cell(row, 2, "Suma h").font = bold
        ws.cell(row, 3, "Soboty").font = bold; ws.cell(row, 4, "Niedziele").font = bold; ws.cell(row, 5, "Święta").font = bold
        row += 1
        for s in summary:
            ws.cell(row, 1, s["employee"]); ws.cell(row, 2, s["hours"])
            ws.cell(row, 3, s["saturdays"]); ws.cell(row, 4, s["sundays"]); ws.cell(row, 5, s["holidays"])
            row += 1

        # --- DODAJ TO TUTAJ ---
        for col_idx in range(1, 3 + len(self.days)):
            col_letter = ws.cell(1, col_idx).column_letter
            # Kolumna 1 i 2 (Pracownik i Typ) potrzebują więcej miejsca
            if col_idx == 1:
                ws.column_dimensions[col_letter].width = 10.72  # Pracownik (trochę szerszy)
            elif col_idx == 2:
                ws.column_dimensions[col_letter].width = 10.5  # Typ danych
            else:
                # Wartość 13.43 w Excelu odpowiada zazwyczaj dokładnie 120 pikselom
                ws.column_dimensions[col_letter].width = 10.3

        wb.save(filename)

    def generate_and_save(self, year, month, employees=None, out_filename=None, initial_stats=None, last_weekend_workers=None, leaves=None):
        # 1. Generujemy dane grafiku (tutaj przekazujemy leaves dalej)
        sched, summ, hol = self.generate(year, month, employees, initial_stats, last_weekend_workers, leaves)

        # 2. Logika unikalnej nazwy pliku
        if out_filename is None:
            base_name = f"harm_{year}_{month:02d}"
            out_filename = f"{base_name}.xlsx"
            
            # Jeśli plik istnieje, dodajemy _v1, _v2, itd.
            version = 1
            while os.path.exists(out_filename):
                out_filename = f"{base_name}_v{version}.xlsx"
                version += 1
        
        # 3. Zapis do Excela
        self.save_xlsx(sched, summ, hol, year, month, out_filename)
        print(f"Sukces! Grafik zapisany jako: {out_filename}")
