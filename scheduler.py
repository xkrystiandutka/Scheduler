import random, calendar
from datetime import datetime, timedelta, date
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from .utils import month_days, polish_holidays
import os


class Scheduler:
    """
    Scheduler z pełnym odbiorem za pracę w soboty/niedziele/święta.
    - 16h przerwy między zmianami
    - Odbiory w dni robocze
    - Weekend: 1 osoba 07-15, 1 osoba 14-22
    - Dni robocze: max 3 osoby na 14-22
    - Sprawiedliwy przydział WS/WN/WP
    """
    MIN_GAP_DAYS = 7
    MAX_SATURDAY = 2
    MAX_SUNDAY = 2
    MAX_HOLIDAY = 2

    def __init__(self, seed=None):
        self.seed = seed
        if seed is not None:
            random.seed(seed)
        self.employees = [
            "F Tomasz","F Krzysztof","S Sławomir","G Artur",
            "K Zbigniew","M Leszek","M Zbigniew","W Magdalena",
            "D Janusz","S Renata"
        ]
        self.SHIFTS = {
            "07-15": (7,15,8),
            "14-22": (14,22,8),
            "08-20": (8,20,12),  # <<< dodane
            "09-21": (9,21,12),  # <<< dodane (opcjonalnie)
            "09-19": (9,19,10),  # <<< dodane (opcjonalnie)
            "OFF": (None,None,0),
            "WN": (None,None,0),
            "WS": (None,None,0),
            "WP": (None,None,0),
            "WW": (None,None,0),
            "WH": (None,None,0),
        }
        self.MIN_SECOND_SHIFT = 1
        self.MAX_SECOND_SHIFT = 3
        self.MIN_REST_HOURS = 16
        self.TARGET_HOURS = 168
        self.COLORS = {
            "saturday":"FF892E","sunday":"FF892E","holiday":"CD3C32","header":"CD3C32"
        }

    def shift_times(self, code, dt):
        if code in ("OFF","WN","WS","WP") or code not in self.SHIFTS:
            return None
        sh, eh, hrs = self.SHIFTS[code]
        start = datetime(dt.year, dt.month, dt.day, sh)
        end = datetime(dt.year, dt.month, dt.day, eh)
        if eh <= sh:
            end += timedelta(days=1)
        return start, end

    def rest_ok(self, prev_code, prev_date, next_code, next_date):
        if prev_code in (None,"OFF","WN","WS","WP") or next_code in (None,"OFF","WN","WS","WP"):
            return True
        prev_times = self.shift_times(prev_code, prev_date)
        next_times = self.shift_times(next_code, next_date)
        if prev_times is None or next_times is None:
            return True
        delta = (next_times[0] - prev_times[1]).total_seconds() / 3600.0
        return delta >= self.MIN_REST_HOURS

    def week_index(self, d):
        first = d.replace(day=1)
        first_monday = first - timedelta(days=first.weekday())
        return (d - first_monday).days // 7

    def _make_weekly_pref(self, weeks, employees):
        weekly_pref = {}
        n = len(employees)
        for w in weeks:
            emp_list = employees.copy()
            random.shuffle(emp_list)
            target_ii = max(1, round(n/3.5))
            weekly_pref[w] = {}
            for i,e in enumerate(emp_list):
                weekly_pref[w][e] = "14-22" if i < target_ii else "07-15"
        return weekly_pref
    
    def consecutive_12h_days(self, schedule, employee, current_day):
        """
        Liczy ile poprzednich dni z rzędu pracownik miał 12h zmiany
        """
        count = 0
        day = current_day - timedelta(days=1)
        while day in schedule[employee] and schedule[employee][day] in ("08-20","09-21"):
            count += 1
            day -= timedelta(days=1)
        return count



    def assign_special_employee(self, employee, year, month, other_employees):
        days = month_days(year, month)
        schedule = {employee: {d: "OFF" for d in days}}
        worked_hours = {employee: 0}
        return schedule, worked_hours
    

    def _assign_weekend_day(self, d, weekly_pref, week_of, schedule, hours, stats, last_sunhol_day, weekend_assigned=None):
        """
        Przydziela zmiany weekendowe i świąteczne:
        - max 2 osoby na dzień
        - żaden pracownik nie może pracować w obu dniach weekendu
        - niedziela/święto: nie więcej niż raz na 7 dni przed lub po
        """
        if weekend_assigned is None:
            weekend_assigned = set()

        weekday = d.weekday()
        is_hol = d in polish_holidays(d.year)

        # Kandydaci: nie pracowali już w tym weekendzie i nie złamali 7-dniowego zakazu
        candidates = []
        for e in schedule.keys():
            if e in weekend_assigned:
                continue
            last = last_sunhol_day.get(e)
            if last and abs((d - last).days) < 7:  # blokada 7 dni przed/po
                continue
            candidates.append(e)

        # jeśli za mało kandydatów, uzupełniamy pozostałych, żeby zawsze były 2 osoby
        if len(candidates) < 2:
            for e in schedule.keys():
                if e not in candidates and e not in weekend_assigned:
                    candidates.append(e)
                if len(candidates) >= 2:
                    break

        # sortowanie kandydatów po statystykach
        candidates.sort(key=lambda e: (stats[e]["weekends"], stats[e]["sundays"], stats[e]["holidays"], random.random()))
        picked = candidates[:2]

        # przydzielamy zmiany
        for i, e in enumerate(picked):
            # blokada dla Renaty 12h
            if e == "S Renata" and self.consecutive_12h_days(schedule, "S Renata", d) < 2:
                schedule[e][d] = "08-20"
                hours[e] += 12
            else:
                if weekday == 5:  # sobota
                    schedule[e][d] = "07-15" if i == 0 else "14-22"
                elif weekday == 6 or is_hol:  # niedziela lub święto
                    schedule[e][d] = "07-15" if i == 0 else "14-22"
            # aktualizacja godzin i statystyk
            shift = schedule[e][d]
            hours[e] += self.SHIFTS[shift][2] if shift not in ("WN","WS","WP","WH") else 0
            stats[e]["weekends"] += 1
            if weekday == 6:
                stats[e]["sundays"] += 1
            if is_hol:
                stats[e]["holidays"] += 1

            last_sunhol_day[e] = d
            weekend_assigned.add(e)



    def _assign_weekday(self, d, weekly_pref, week_of, schedule, hours, stats):
        w = week_of[d]
        cand = list(weekly_pref[w].keys())
        random.shuffle(cand)
        desired_ii = [e for e in cand if weekly_pref[w][e]=="14-22"]
        if len(desired_ii) > self.MAX_SECOND_SHIFT:
            desired_ii = sorted(desired_ii, key=lambda x:(stats[x]["weekends"], stats[x]["sundays"]))[:self.MAX_SECOND_SHIFT]

        for e in cand:
            if e == "S Renata":   # 🔴 KLUCZOWE
                continue
            if schedule[e][d] == "WW":  # <-- ochrona przed nadpisaniem
                continue
            target = "14-22" if e in desired_ii else "07-15"
            prev = d - timedelta(days=1)
            prev_shift = schedule[e].get(prev, "OFF")
            if not self.rest_ok(prev_shift, prev, target, d):
                if schedule[e][d] == "WW":  # <-- ochrona przed nadpisaniem
                    continue
                other = "07-15" if target=="14-22" else "14-22"
                if self.rest_ok(prev_shift, prev, other, d):
                    if schedule[e][d] == "WW":  # <-- ochrona przed nadpisaniem
                        continue
                    schedule[e][d] = other
                    hours[e] += self.SHIFTS[other][2]
                else:
                    schedule[e][d] = "OFF"
            else:
                schedule[e][d] = target
                hours[e] += self.SHIFTS[target][2]


   
    def _assign_compensatory(self, employees, days, schedule, hours, week_of, last_sunhol_day):
        holidays = set(polish_holidays(days[0].year))

        # dni robocze do odbiorów
        workdays = [
            d for d in days
            if d.weekday() < 5 and d not in holidays
        ]

        # ile odbiorów jest w danym dniu
        day_load = Counter()

        for e in employees:
            for d in days:
                shift = schedule[e].get(d)
                if shift not in ("07-15", "14-22"):
                    continue

                # typ odbioru
                if e == "S Renata" and shift == "08-20" and d.weekday() < 5:
                    comp = "WH"
                elif d.weekday() == 6:        # niedziela
                    comp = "WN"
                elif d in holidays:         # święto
                    comp = "WS"
                elif d.weekday() == 5:      # sobota
                    comp = "WP"
                else:
                    continue

                # możliwe dni odbioru
                possible = [
                    wd for wd in workdays
                    if wd not in holidays
                    and wd.weekday() < 5
                    and schedule[e].get(wd) in ("OFF", "07-15", "14-22")
                    and abs((wd - d).days) <= 7                           # max 7 dni przed lub po

                ]


                if not possible:
                    continue

                # wybór dnia: najmniej odbiorów
                possible.sort(
                    key=lambda wd: (day_load[wd], random.random())
                )

                cd = possible[0]
                schedule[e][cd] = comp
                day_load[cd] += 1



    def _balance_hours(self, employees, days, schedule, hours):
        target = self.TARGET_HOURS
        max_iters = 5000
        it = 0
        improved = True

        while it < max_iters and improved:
            it += 1
            improved = False

            over = sorted(
                [e for e in employees if hours[e] > target],
                key=lambda x: hours[x] - target,
                reverse=True
            )
            under = sorted(
                [e for e in employees if hours[e] < target],
                key=lambda x: target - hours[x],
                reverse=True
            )

            if not over or not under:
                break

            a = over[0]
            b = under[0]

            # ❌ NIE BALANSUJEMY RENATY
            if a == "S Renata" or b == "S Renata":
                break

            random.shuffle(days)

            for d in days:
                sa = schedule[a].get(d)
                sb = schedule[b].get(d)

                if sa in ("07-15","14-22") and sb == "OFF":
                    prev_b = d - timedelta(days=1)
                    if not self.rest_ok(schedule[b].get(prev_b,"OFF"), prev_b, sa, d):
                        continue

                    ha = hours[a] - self.SHIFTS[sa][2]
                    hb = hours[b] + self.SHIFTS[sa][2]

                    if abs(ha - target) < abs(hours[a] - target) and \
                    abs(hb - target) < abs(hours[b] - target):

                        schedule[b][d] = sa
                        schedule[a][d] = "OFF"
                        hours[a] = ha
                        hours[b] = hb
                        improved = True
                        break

                if sb in ("07-15","14-22") and sa == "OFF":
                    prev_a = d - timedelta(days=1)
                    if not self.rest_ok(schedule[a].get(prev_a,"OFF"), prev_a, sb, d):
                        continue

                    ha = hours[a] + self.SHIFTS[sb][2]
                    hb = hours[b] - self.SHIFTS[sb][2]

                    if abs(ha - target) < abs(hours[a] - target) and \
                    abs(hb - target) < abs(hours[b] - target):

                        schedule[a][d] = sb
                        schedule[b][d] = "OFF"
                        hours[a] = ha
                        hours[b] = hb
                        improved = True
                        break


    def _assign_end_month(self, days, schedule, hours, stats):
        """
        UZUPEŁNIA obsadę ostatnich dni miesiąca (weekend / święto),
        ale:
        - NIGDY nie przekracza 2 osób
        - NIE nadpisuje istniejących zmian
        - NIE dokłada ludzi, jeśli już są 2
        """

        holidays = set(polish_holidays(days[0].year))

        # bierzemy tylko OSTATNI dzień miesiąca
        end_days = [days[-1]]

        for d in end_days:
            weekday = d.weekday()
            is_hol = d in holidays

            # interesują nas tylko sobota / niedziela / święto
            if weekday < 5 and not is_hol:
                continue

            # 🔴 KROK 1: sprawdzamy kto JUŻ pracuje
            working = [
                e for e in schedule
                if schedule[e][d] in ("07-15", "14-22", "08-20")
            ]

            # 🔴 KROK 2: jeśli już są 2 osoby → NIC NIE ROBIMY
            if len(working) >= 2:
                continue

            # 🔴 KROK 3: ustalamy jakie zmiany są już zajęte
            assigned_shifts = [schedule[e][d] for e in working]

            # 🔴 KROK 4: kandydaci tylko z OFF
            candidates = sorted(
                [e for e in schedule if schedule[e][d] == "OFF"],
                key=lambda e: (
                    stats[e]["weekends"],
                    stats[e]["sundays"],
                    stats[e]["holidays"],
                    random.random()
                )
            )

            # 🔴 KROK 5: uzupełniamy TYLKO do 2 osób
            for e in candidates:
                if len(assigned_shifts) >= 2:
                    break

                # wybór zmiany – brakująca
                shift = "07-15" if "07-15" not in assigned_shifts else "14-22"

                schedule[e][d] = shift
                hours[e] += self.SHIFTS[shift][2]

                stats[e]["weekends"] += 1
                if weekday == 6:
                    stats[e]["sundays"] += 1
                if is_hol:
                    stats[e]["holidays"] += 1

                assigned_shifts.append(shift)

    def _assign_renata_weekdays(self, days, schedule, hours, holidays):
        workdays = [d for d in days if d.weekday() < 5 and d not in holidays]
        target = self.renata_target_hours

        # 1️⃣ policz poprawną kombinację
        n_12h = target // 12
        rest = target - n_12h * 12

        if rest == 4:
            n_12h -= 1
            rest += 12

        n_8h = rest // 8

        # wyczyść dni robocze
        for d in workdays:
            schedule["S Renata"][d] = "OFF"

        last_12h = None

        # 2️⃣ przydziel zmiany
        for d in workdays:
            if n_12h > 0:
                if last_12h and (d - last_12h).days == 1:
                    if n_8h > 0:
                        schedule["S Renata"][d] = "07-15"
                        hours["S Renata"] += 8
                        n_8h -= 1
                        last_12h = None
                else:
                    schedule["S Renata"][d] = "08-20"
                    hours["S Renata"] += 12
                    n_12h -= 1
                    last_12h = d

            elif n_8h > 0:
                schedule["S Renata"][d] = "07-15"
                hours["S Renata"] += 8
                n_8h -= 1

        # 3️⃣ reszta dni roboczych = WH
        for d in workdays:
            if schedule["S Renata"][d] == "OFF":
                schedule["S Renata"][d] = "WH"


    def _assign_wh_days(self, schedule, employee, days, holidays):
        target = self.renata_target_hours

        def current_hours():
            return sum(
                self.SHIFTS.get(schedule[employee][d], (0,0,0))[2]
                for d in days
            )

        excess = current_hours() - target
        if excess <= 0:
            print(f">>> {employee} - brak nadmiaru godzin, nie dodajemy WH")
            return

        # kandydaci: dni robocze z pracą
        candidates = [
            d for d in days
            if d.weekday() < 5
            and d not in holidays
            and schedule[employee][d] in ("08-20", "07-15", "14-22")
        ]

        # losowo tasujemy
        random.shuffle(candidates)

        # sortujemy: najpierw 12h, potem 8h
        candidates.sort(key=lambda d: -self.SHIFTS[schedule[employee][d]][2])

        wh_per_week = defaultdict(int)
        wh_assigned = []

        for d in candidates:
            if excess <= 0:
                break

            # sprawdzamy, czy poprzedni dzień nie jest WH
            prev_day = d - timedelta(days=1)
            next_day = d + timedelta(days=1)
            if (prev_day in schedule[employee] and schedule[employee][prev_day] == "WH") or \
            (next_day in schedule[employee] and schedule[employee][next_day] == "WH"):
                continue  # pomijamy, żeby mieć przerwę 1 dnia

            w = self.week_index(d)
            if wh_per_week[w] >= 2:
                continue

            hrs = self.SHIFTS[schedule[employee][d]][2]

            schedule[employee][d] = "WH"
            wh_per_week[w] += 1
            wh_assigned.append(d)
            excess -= hrs



    def generate(self, year, month, employees=None):
        if employees is None:
            employees = self.employees.copy()
        days = month_days(year, month)
        holidays = polish_holidays(year)

        workdays = [
            d for d in days
            if d.weekday() < 5 and d not in holidays
        ]

        # teraz target Renaty
        self.renata_target_hours = len(workdays) * 8


        schedule = {e:{d:"OFF" for d in days} for e in employees}
        hours = {e:0 for e in employees}
        week_of = {d:self.week_index(d) for d in days}
        weeks = sorted(set(week_of.values()))
        last_sunhol_week = {e: None for e in employees}
        stats = {e:{"weekends":0,"sundays":0,"holidays":0} for e in employees}

        weekly_pref = self._make_weekly_pref(weeks, employees)

        # 1️⃣ Przypisanie weekendów i świąt
        for d in days:
            weekday = d.weekday()
            if weekday in (5,6) or d in holidays:
                self._assign_weekend_day(d, weekly_pref, week_of, schedule, hours, stats, last_sunhol_week)
            else:
                self._assign_weekday(d, weekly_pref, week_of, schedule, hours, stats)

        self._assign_renata_weekdays(days, schedule, hours, holidays)
        holidays = set(polish_holidays(year))


        # 2️⃣ Ostatni dzień miesiąca: zapewnienie 2 osób, przydzielenie wcześniejszych odbiorów
        self._assign_end_month(days, schedule, hours, stats)

        # 3️⃣ Kompensatory za wszystkie weekendy i święta
        self._assign_compensatory(employees, days, schedule, hours, week_of, last_sunhol_week)
        self._assign_wh_days(schedule, "S Renata", days, holidays)

        # 4️⃣ Podsumowanie godzin
        working_shifts = ("07-15","14-22","08-20")  # wszystkie zmiany normalne


        summary = []
        for e in employees:
            hrs = sum(
                0 if schedule[e][d] in ("OFF","WN","WS","WP", "WW", "WH") else self.SHIFTS[schedule[e][d]][2]
                for d in schedule[e]
            )
            summary.append({
                "employee": e,
                "hours": hrs,
                "weekends": sum(1 for d in schedule[e] if d.weekday() >= 5 and schedule[e][d] in working_shifts),
                "sundays": sum(1 for d in schedule[e] if d.weekday() == 6 and schedule[e][d] in working_shifts),
                "holidays": sum(1 for d in schedule[e] if d in polish_holidays(d.year) and schedule[e][d] in working_shifts),
            })

        return schedule, summary, holidays


    def save_xlsx(self, schedule, summary, holidays, year, month, filename=None):
        if filename is None:
            filename = f"harm_{year}_{month:02d}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = f"{calendar.month_name[month]}_{year}"
        days = sorted(next(iter(schedule.values())).keys())
        nd = len(days)

        border = Border(left=Side(style="thin"),right=Side(style="thin"),
                        top=Side(style="thin"),bottom=Side(style="thin"))
        center = Alignment(horizontal="center",vertical="center")
        bold = Font(bold=True)
        sat_fill = PatternFill("solid", fgColor=self.COLORS["saturday"])
        sun_fill = PatternFill("solid", fgColor=self.COLORS["sunday"])
        hol_fill = PatternFill("solid", fgColor=self.COLORS["holiday"])
        hdr_fill = PatternFill("solid", fgColor=self.COLORS["header"])
        odb_fill = PatternFill("solid", fgColor="92D050")  # zielony dla odbiorów

        # nagłówki
        ws.cell(row=1,column=1,value="Nr ewidenc. pracowniKa").font=bold
        ws.cell(row=1,column=2,value="Pracownik").font=bold
        ws.cell(row=1,column=3,value="Godz. pracy od - do").font=bold

        for i,d in enumerate(days):
            c = 4+i
            ws.cell(row=1,column=c,value=d.day).font=bold
            ws.cell(row=2,column=c,value=calendar.day_name[d.weekday()][:2]).font=bold
            if d in holidays:
                ws.cell(row=1,column=c).fill = hol_fill; ws.cell(row=2,column=c).fill=hol_fill
            elif d.weekday()==5:
                ws.cell(row=1,column=c).fill = sat_fill; ws.cell(row=2,column=c).fill=sat_fill
            elif d.weekday()==6:
                ws.cell(row=1,column=c).fill = sun_fill; ws.cell(row=2,column=c).fill=sun_fill

        row=3
        for e,days_map in schedule.items():
            ws.cell(row=row,column=2,value=e).font=bold
            ws.cell(row=row,column=3,value="Godz. pracy od - do").font=bold
            for i,d in enumerate(days):
                c=4+i
                val = days_map[d]
                ws.cell(row=row,column=c,value=val)
                # kolorowanie
                if val in ("WN","WS","WP", "WH"):
                    ws.cell(row=row,column=c).fill = odb_fill
                elif d in holidays:
                    ws.cell(row=row,column=c).fill=hol_fill
                elif d.weekday()==5:
                    ws.cell(row=row,column=c).fill=sat_fill
                elif d.weekday()==6:
                    ws.cell(row=row,column=c).fill=sun_fill
                ws.cell(row=row,column=c).alignment = center
                ws.cell(row=row,column=c).border = border
            row+=1
            # godziny dla każdego dnia
            ws.cell(row=row,column=3,value="Liczba godz.").font=bold
            for i,d in enumerate(days):
                c=4+i
                v = days_map[d]
                if v in ("WN","WS","WP", "WW", "WH"):
                    hrs = 0  # odbiór = 8h
                elif v == "OFF":
                    hrs = 0  # dzień wolny = 0h
                else:
                    hrs = self.SHIFTS.get(v,(None,None,0))[2]
                ws.cell(row=row,column=c,value=hrs)
                ws.cell(row=row,column=c).alignment = center
                ws.cell(row=row,column=c).border = border
            row+=1

        row+=1
        ws.cell(row=row,column=1,value="PODSUMOWANIE").font=Font(bold=True,underline="single"); row+=1
        ws.cell(row=row,column=1,value="Pracownik").font=bold
        ws.cell(row=row,column=2,value="Godziny").font=bold
        ws.cell(row=row,column=3,value="Weekend").font=bold
        ws.cell(row=row,column=4,value="Niedziela").font=bold
        ws.cell(row=row,column=5,value="Święta").font=bold
        row+=1
        smap = {s["employee"]:(s["hours"],s["weekends"],s["sundays"],s["holidays"]) for s in summary}
        for e in schedule.keys():
            vals = smap.get(e,(0,0,0,0))
            ws.cell(row=row,column=1,value=e)
            ws.cell(row=row,column=2,value=vals[0])
            ws.cell(row=row,column=3,value=vals[1])
            ws.cell(row=row,column=4,value=vals[2])
            ws.cell(row=row,column=5,value=vals[3])
            row+=1

        wb.save(filename)
        print("Saved:", filename)        

    def generate_and_save(self, year, month, employees=None, out_filename=None):
        sched, summ, hol = self.generate(year, month, employees)

        if out_filename is None:
            base_name = f"harm_{year}_{month:02d}"
            out_filename = f"{base_name}.xlsx"
            
            # jeśli plik istnieje, dodajemy _v1, _v2, ...
            version = 1
            while os.path.exists(out_filename):
                out_filename = f"{base_name}_v{version}.xlsx"
                version += 1

        self.save_xlsx(sched, summ, hol, year, month, out_filename)